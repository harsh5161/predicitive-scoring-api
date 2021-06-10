import pandas as pd
from openpyxl import load_workbook
import csv
import numpy as np
import collections


def importFile(nrows, sheet_name, path):

    print('#### RUNNING WAIT ####')

    # IF THE EXTENSION IS CSV
    def importCsv(path):

        print('We have a csv file')
        try:
            df = pd.read_csv(path, low_memory=False,
                             nrows=nrows, error_bad_lines=False)
            if df.shape[1] == 1:
                df = pd.read_csv(path, low_memory=False, sep=';', nrows=nrows)
            # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df

        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

        except UnicodeDecodeError:
            try:
                enc = 'unicode_escape'
                df = pd.read_csv(path, encoding=enc, low_memory=False,
                                 nrows=nrows, error_bad_lines=False)
                # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                return df

            except UnicodeDecodeError:
                try:
                    enc = 'ISO-8859-1'
                    df = pd.read_csv(
                        path, encoding=enc, low_memory=False, nrows=nrows, error_bad_lines=False)
                    # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                    return df
                except:
                    pass

        except:
            try:
                df = pd.read_csv(path, nrows=nrows, error_bad_lines=False)
                separators = ["~", "!", "@", "#", "$", "%", "^", "&",
                              "*", ":", "|", "/", ";"]     # all possible separators
                # if separator was "," we would have more than 1 columns
                if len(df.columns) <= 3:
                    cols = df.columns[0]
                    possibleSep = []
                    for i in separators:                                    # checking all the separators present in column names
                        if i in cols:
                            possibleSep.append(i)

                    # iterate through possible seprators till we get the correct one
                    for j in possibleSep:
                        df_sep = pd.read_csv(
                            path, sep=j, nrows=nrows, error_bad_lines=False)
                        if len(df_sep.columns) > 3:
                            # print('This file has {} columns and {} rows'.format(df_sep.shape[1],df_sep.shape[0]))
                            return df_sep
            except:
                try:
                    # for tab ie "\" tsv files
                    if len(pd.read_csv(path, sep=None).columns, nrows=nrows, error_bad_lines=False) > 3:
                        df = pd.read_csv(
                            path, sep=None, nrows=nrows, error_bad_lines=False)
                        # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                        return df
                except:
                    pass

    # IF THE EXTENSION IS JSON
    def importJSON(path):
        try:
            print('We have a JSON file')
            df = pd.read_json(path)
            # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df
        except Exception:
            try:
                df = pd.read_json(path, lines=True)
                # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
                return df

            except ValueError:
                print('File not found, Check the name, path, spelling mistakes')
                error = True
                return None

    # IF THE EXTENSION IS XL
    def importExcel(sheet_name, path):
        try:
            print('We have an Excel file')
            #######
            # opening workbook
            wb = load_workbook(path)

            if len(wb.sheetnames) == 1:
                data = wb[wb.sheetnames[0]].values
                cols = next(data)[0:]
                sheet_name = None
            else:
                if sheet_name:
                    val = sheet_name
                else:
                    val = input('Input the name of the sheet')
                data = wb[val].values
                cols = next(data)[0:]

            df = pd.DataFrame(data, columns=cols)
            print(df.head(10))
            return df, val

            #######
        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

    def importTable(path):
        try:
            print('We have General Table File')
            df = pd.read_table(path, nrows=nrows)
            if df.shape[1] == 1:
                df = pd.read_table(path, sep=',', nrows=nrows)
            # print('This file has {} columns and {} rows'.format(df.shape[1],df.shape[0]))
            return df
        except FileNotFoundError:
            print('File not found, Check the name, path, spelling mistakes')
            error = True
            return None

    try:
        ext = path.split('.')[1].lower().strip()
        print('extension is {}'.format(ext))
        if ext == 'csv' or ext == 'tsv':
            df = importCsv(path)
            df = duplicateHandler(df)
            return df, None
        elif ext == 'json':
            df = importJSON(path)
            df = duplicateHandler(df)
            return df, None
        elif 'xl' in ext:
            df, sheet_name = importExcel(sheet_name, path)
            return df, sheet_name
        elif ext == 'data':
            df = importTable(path)
            df = duplicateHandler(df)
            return df, None
        else:
            print('File format not supported\n')
    except Exception as e:
        print('We ran into some Error!')
        print('The error message is {}'.format(e))
        return None, None


def dataHandler(dx, target=None):
    update = False
    for col in dx.columns:  # Counting the non-null values present in a column and removing them if necessary
        if 'Unnamed' in col:
            if dx[col].count() < 0.5*dx.shape[0]:
                dx.drop(col, axis=1, inplace=True)
                update = True

    # checking if there are string specified null values in the target column and replacing it
    if target is not None:
        if str(dx[target].dtype) == 'object':
            dx[target].replace({'NA': np.nan}, inplace=True)
            if dx[target].nunique() > 5:
                dx[target] = dx[target].apply(
                    lambda x: removeSpecialCharacters(x))
                try:
                    dx[target] = pd.to_numeric(dx[target])
                    if str(dx[target].dtype) != 'object':
                        dx[target] = dx[target].astype(float)
                except Exception as e:
                    print(f'Exception has occurred : {e}')

    # to handel cases when some blank rows or other information above the data table gets assumed to be column name
    # Checking for unnamed columns
    if (len([col for col in dx.columns if 'Unnamed' in col]) > 0.5*dx.shape[1]):
        # Getting the values in the first row of the dataframe into a list
        colNew = dx.loc[0].values.tolist()
        dx.columns = colNew  # Making values stored in colNew as the new column names
        # dropping the row whose values we made as the column names
        dx = dx.drop(labels=[0])
        # resetting index to the normal pattern 0,1,2,3...
        dx.reset_index(drop=True, inplace=True)
    else:
        return dx, update

    # Following three lines of code are for counting the number of null values in our new set of column names
    new_column_names = dx.columns.values.tolist()
    new_column_names = pd.DataFrame(new_column_names)
    null_value_sum = new_column_names.isnull().sum()[0]
    # if count of null values are less than a certain ratio of total no of columns
    if null_value_sum < 0.5*dx.shape[1]:
        return dx, update
    while(null_value_sum >= 0.5*dx.shape[1]):
        colNew = dx.loc[0].values.tolist()
        dx.columns = colNew
        dx = dx.drop(labels=[0])
        dx.reset_index(drop=True, inplace=True)
        new_column_names = dx.columns.values.tolist()
        new_column_names = pd.DataFrame(new_column_names)
        null_value_sum = new_column_names.isnull().sum()[0]

    return dx, update


def duplicateHandler(df):

    # dealing with column names that are empty strings
    df.rename(columns={"": 'Unnamed'}, inplace=True)
    actual = df.columns.to_list()
    a = [x.strip().lower() for x in df.columns.to_list()]
    dups = [item for item, count in collections.Counter(
        a).items() if count > 1]

    for i in range(len(a)):
        if a[i] in dups:
            actual[i] = f'{actual[i].strip()}_{i}'

    df.columns = actual

    return df


def getDF(df, model):
    try:
        mdf = df[model['init_cols'].drop(model['Target'])]
        print('Columns Match!')
        return mdf
    except KeyError as e:
        print('We could not find the column/columns ' +
              str(e) + ' in the current file!')
        print(
            'The column names don\'t match with the ones that were present during Training')
        print('Kindly Check for spelling, upper/lower cases and missing columns if any!')
        return None


def removeSpecialCharacters(x):
    flag = 0
    temp = str(x).lower()

    if '$' in str(x)[0]:
        flag = 1
    elif '€' in str(x)[0]:
        flag = 1
    elif 'k' in temp[-1]:
        flag = 1
    elif 'm' in temp[-1]:
        flag = 1
    elif 'b' in temp[-1]:
        flag = 1
    elif ',' in temp:
        flag = 1

    if flag == 1:
        try:
            temp = temp.replace('€', '')
        except:
            pass
        try:
            temp = temp.replace('k', '000')
        except:
            pass
        try:
            temp = temp.replace('m', '000000')
        except:
            pass
        try:
            temp = temp.replace('b', '000000000')
        except:
            pass
        try:
            temp = temp.replace('$', '')
        except:
            pass
        try:
            temp = temp.replace(',', '')
        except:
            pass
        try:
            return pd.to_numeric(temp)
        except:
            return temp
    else:
        return x
